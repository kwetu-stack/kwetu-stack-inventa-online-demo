import io
import os
import re
import secrets
import sqlite3

import qrcode
from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import timedelta
from flask import session
import time


# ---------- APP CONFIG ----------
load_dotenv()
app = Flask(__name__)
app.permanent_session_lifetime = timedelta(minutes=15)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"]
)


def load_secret_key():
    secret_key = os.environ.get("SECRET_KEY")
    if secret_key:
        return secret_key

    env_path = ".env"
    if not os.path.exists(env_path):
        return None

    try:
        with open(env_path, "r", encoding="utf-8") as env_file:
            for raw_line in env_file:
                line = raw_line.strip()
                if not line or line.startswith("#"):
                    continue

                if line.startswith("SECRET_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")

                if line.startswith("$env:SECRET_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    except OSError:
        return None

    return None


app.secret_key = load_secret_key()

if not app.secret_key:
    raise RuntimeError("SECRET_KEY is not set. Please configure environment variable.")

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

TENANT = "digitalclub"
DB_PATH = "inventory.db"
DEMO_MODE = os.environ.get("DEMO_MODE", "true").strip().lower() in {"1", "true", "yes", "on"}
DEMO_WRITE_BLOCKED_ENDPOINTS = {
    "change_password",
    "add_stock",
    "record_sale",
    "upload_inventory",
}

#----Helpers-----


def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def redirect_to_tenant_path(path):
    return redirect(f"/{TENANT}/{path}")


def parse_positive_int(value):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def sanitize_text(value, max_length=255):
    if not value:
        return ""

    value = value.strip()

    # Remove script tags completely
    value = re.sub(r"<.*?>", "", value)

    # Remove remaining dangerous characters
    value = re.sub(r"[\"';]", "", value)

    return value[:max_length]


def redirect_back(default_endpoint="dashboard"):
    return redirect(request.referrer or url_for(default_endpoint))


#---Decorator------
def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if DEMO_MODE:
            return view_func(*args, **kwargs)

        if "user_id" not in session:
            flash("Please log in to continue.", "error")
            return redirect(url_for("login"))

        conn = get_db_connection()
        user = conn.execute(
            "SELECT session_token FROM users WHERE id = ?",
            (session["user_id"],)
        ).fetchone()
        conn.close()

        if not user or user["session_token"] != session.get("session_token"):
            session.clear()
            flash("Session expired. Please log in again.", "error")
            return redirect(url_for("login"))

        return view_func(*args, **kwargs)

    return wrapped_view


def role_required(required_role):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(*args, **kwargs):
            if DEMO_MODE:
                return view_func(*args, **kwargs)

            if "user_id" not in session:
                flash("Please log in to continue.", "error")
                return redirect(url_for("login"))

            if session.get("role") != required_role:
                return render_template("403.html", tenant=TENANT), 403

            return view_func(*args, **kwargs)

        return wrapped_view

    return decorator

# ---------- CSRF PROTECTION ----------
def generate_csrf_token():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(16)
    return session["_csrf_token"]


app.jinja_env.globals["csrf_token"] = generate_csrf_token


def validate_csrf():
    token = session.get("_csrf_token")
    form_token = request.form.get("csrf_token")

    if not token or not form_token or token != form_token:
        flash("Invalid CSRF token.", "error")
        return False
    return True


@app.context_processor
def inject_demo_mode():
    return {"demo_mode": DEMO_MODE}


def log_action(user_id, action, details=""):
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO logs (user_id, action, details) VALUES (?, ?, ?)",
            (user_id, action, details),
        )
        conn.commit()
        conn.close()
    except sqlite3.Error:
        pass  # Logging should never crash the app


@app.before_request
def block_suspicious_input():
    suspicious_patterns = ["<script", "DROP TABLE", "--", ";--"]

    for value in request.values.values():
        if isinstance(value, str):
            for pattern in suspicious_patterns:
                if pattern.lower() in value.lower():
                    flash("Invalid or unsafe input detected.", "error")
                    return redirect(request.referrer or url_for("login"))


@app.before_request
def enforce_demo_mode_read_only():
    if not DEMO_MODE:
        return None

    if request.method == "POST" and request.endpoint in DEMO_WRITE_BLOCKED_ENDPOINTS:
        flash(
            "Demo mode is active. You can explore every page, but changes are not saved.",
            "info",
        )
        return redirect_back()

    return None

@app.before_request
def session_timeout_check():
    if DEMO_MODE:
        return None

    if "user_id" in session:
        now = int(time.time())
        last_activity = session.get("last_activity")

        if last_activity:
            if now - last_activity > 900:  # 15 minutes
                session.clear()
                flash("Session expired. Please log in again.", "error")
                return redirect(url_for("login"))

        session["last_activity"] = now                

# ---------- LOGIN ----------

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute")
def login():
    if DEMO_MODE:
        return redirect_to_tenant_path("dashboard")

    if request.method == "POST":
        if not validate_csrf():
            return redirect(url_for("login"))
        username = sanitize_text(request.form.get("username"))
        password = sanitize_text(request.form.get("password"))

        if not username or not password:
            flash("Username and password are required.", "error")
            return redirect(url_for("login"))

        try:
            conn = get_db_connection()
            user = conn.execute(
                "SELECT * FROM users WHERE username = ?",
                (username,),
            ).fetchone()
            conn.close()
        except sqlite3.Error:
            flash("Database error while logging in. Please try again.", "error")
            return redirect(url_for("login"))

        if user and check_password_hash(user["password"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            session["last_activity"] = int(time.time())

            conn = get_db_connection()
            conn.execute(
                "UPDATE users SET failed_attempts = 0, locked_until = 0 WHERE id = ?",
                (user["id"],),
            )
            import uuid

            token = str(uuid.uuid4())

            conn.execute(
                "UPDATE users SET session_token = ? WHERE id = ?",
                (token, user["id"]),
            )

            session["session_token"] = token
            conn.commit()
            conn.close()

            log_action(user["id"], "LOGIN", "User logged in")
            return redirect_to_tenant_path("dashboard")

        flash("Invalid username or password.", "error")
        return redirect(url_for("login"))

    return render_template("login.html", tenant=TENANT)


# ---------- LOGOUT ----------
@app.route("/logout")
@login_required
def logout():
    if DEMO_MODE:
        return redirect_to_tenant_path("dashboard")

    if "user_id" in session:
        conn = get_db_connection()
        conn.execute(
            "UPDATE users SET session_token = NULL WHERE id = ?",
            (session["user_id"],)
        )
        conn.commit()
        conn.close()

    session.clear()
    flash("Logged out successfully.", "success")
    return redirect(url_for("login"))


# ---------- CHANGE PASSWORD (REDIRECT) ----------
@app.route("/change-password")
@login_required
def change_password_redirect():
    if DEMO_MODE:
        return redirect_to_tenant_path("dashboard")

    return redirect_to_tenant_path("change-password")


# ---------- CHANGE PASSWORD ----------
@app.route(f"/{TENANT}/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if DEMO_MODE:
        return redirect_to_tenant_path("dashboard")

    if request.method == "POST":
        if not validate_csrf():
            return redirect_to_tenant_path("change-password")
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()

        if not current_password or not new_password:
            flash("All fields are required.", "error")
            return redirect_to_tenant_path("change-password")

        if len(new_password) < 6:
            flash("New password must be at least 6 characters.", "error")
            return redirect_to_tenant_path("change-password")

        try:
            conn = get_db_connection()
            user = conn.execute(
                "SELECT * FROM users WHERE id = ?",
                (session["user_id"],),
            ).fetchone()

            if user and check_password_hash(user["password"], current_password):
                hashed_password = generate_password_hash(new_password)
                conn.execute(
                    "UPDATE users SET password = ? WHERE id = ?",
                    (hashed_password, session["user_id"]),
                )
                conn.commit()
                flash("Password updated successfully!", "info")
            else:
                flash("Current password is incorrect.", "error")

            conn.close()

        except sqlite3.Error:
            flash("Database error while updating password.", "error")

        return redirect_to_tenant_path("dashboard")

    return render_template("change-password.html", tenant=TENANT)

# ---------- DASHBOARD ----------
@app.route("/")
def home():
    return redirect_to_tenant_path("dashboard")


@app.route(f"/{TENANT}/dashboard")
@login_required
def dashboard():
    try:
        conn = get_db_connection()

        products = conn.execute(
            "SELECT * FROM products WHERE tenant = ?",
            (TENANT,),
        ).fetchall()

        total_products = len(products)

        total_sales = conn.execute(
            "SELECT SUM(grand_total) FROM sales WHERE tenant = ?",
            (TENANT,),
        ).fetchone()[0] or 0

        sales_data = conn.execute(
            "SELECT date, SUM(grand_total) as total FROM sales WHERE tenant = ? GROUP BY date ORDER BY date",
            (TENANT,),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to load dashboard data right now.", "error")
        total_products = 0
        total_sales = 0
        sales_data = []

    dates = [row["date"] for row in sales_data]
    totals = [row["total"] for row in sales_data]

    return render_template(
        "dashboard.html",
        tenant=TENANT,
        total_products=total_products,
        total_sales=total_sales,
        dates=dates,
        totals=totals,
    )

# ---------- VIEW PRODUCTS ----------
@app.route(f"/{TENANT}/products")
@login_required
def view_products():
    try:
        conn = get_db_connection()

        products = conn.execute(
            "SELECT * FROM products WHERE tenant = ?",
            (TENANT,),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to load products right now.", "error")
        products = []

    return render_template(
        "view-products.html",
        products=products,
        tenant=TENANT,
    )


# ---------- ADD STOCK ----------
@app.route(f"/{TENANT}/add-stock", methods=["GET", "POST"])
@login_required
@role_required("admin")
def add_stock():

    if request.method == "POST":
        if not validate_csrf():
            return redirect_to_tenant_path("add-stock")
        product_id = parse_positive_int(request.form.get("product_id"))
        quantity = parse_positive_int(request.form.get("quantity"))

        if not product_id or not quantity:
            flash("Please select a valid product and quantity.", "error")
            return redirect_to_tenant_path("add-stock")

        try:
            conn = get_db_connection()

            updated = conn.execute(
                "UPDATE products SET quantity = quantity + ? WHERE id = ? AND tenant = ?",
                (quantity, product_id, TENANT),
            )

            conn.commit()

            # Log stock update
            conn.execute(
                "INSERT INTO logs (user_id, action, details) VALUES (?, ?, ?)",
                (
                    session["user_id"],
                    "STOCK_UPDATE",
                    f"Added {quantity} units to product ID {product_id}",
                ),
            )

            conn.commit()
            conn.close()

        except sqlite3.Error:
            flash("Unable to update stock right now.", "error")
            return redirect_to_tenant_path("add-stock")

        if updated.rowcount == 0:
            flash("Selected product was not found.", "error")
        else:
            flash("Stock added successfully.", "info")

        return redirect_to_tenant_path("add-stock")

    try:
        conn = get_db_connection()

        products = conn.execute(
            "SELECT id, name FROM products WHERE tenant = ?",
            (TENANT,),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to load products for stock update.", "error")
        products = []

    return render_template("add-stock.html", products=products, tenant=TENANT)


# ---------- RECORD SALE ----------
@app.route(f"/{TENANT}/record-sale", methods=["GET", "POST"])
@limiter.limit("20 per minute")
@login_required
def record_sale():

    try:
        conn = get_db_connection()
        products = conn.execute(
            "SELECT * FROM products WHERE tenant = ?",
            (TENANT,),
        ).fetchall()
    except sqlite3.Error:
        flash("Unable to load sale form right now.", "error")
        return redirect_to_tenant_path("dashboard")

    if request.method == "POST":
        if not validate_csrf():
            return redirect_to_tenant_path("record-sale")
        customer = sanitize_text(request.form.get("customer"))
        product_ids = request.form.getlist("product_id[]")
        quantities = request.form.getlist("quantity[]")

        if not customer:
            conn.close()
            flash("Customer name is required.", "error")
            return redirect_to_tenant_path("record-sale")

        if not product_ids or len(product_ids) != len(quantities):
            conn.close()
            flash("Please provide valid sale items.", "error")
            return redirect_to_tenant_path("record-sale")

        total_amount = 0
        selected_items = []

        for index, product_id in enumerate(product_ids):
            qty = parse_positive_int(quantities[index])
            parsed_product_id = parse_positive_int(product_id)

            if not qty or not parsed_product_id:
                conn.close()
                flash("Each sale item must have a valid product and quantity.", "error")
                return redirect_to_tenant_path("record-sale")

            product = conn.execute(
                "SELECT * FROM products WHERE id = ? AND tenant = ?",
                (parsed_product_id, TENANT),
            ).fetchone()
            if not product:
                conn.close()
                flash("One of the selected products no longer exists.", "error")
                return redirect_to_tenant_path("record-sale")

            if product["quantity"] < qty:
                conn.close()
                flash(f"Not enough stock for {product['name']}.", "error")
                return redirect_to_tenant_path("record-sale")

            total = qty * product["price"]
            total_amount += total
            selected_items.append((parsed_product_id, qty, product["price"]))

        if not selected_items:
            conn.close()
            flash("Please add at least one product to the sale.", "error")
            return redirect_to_tenant_path("record-sale")

        vat = round(total_amount * 0.16, 2)
        grand_total = round(total_amount + vat, 2)

        try:
            cursor = conn.cursor()
            cursor.execute(
                'INSERT INTO sales (customer, total_amount, vat, grand_total, date, tenant) VALUES (?, ?, ?, ?, DATE("now"), ?)',
                (customer, total_amount, vat, grand_total, TENANT),
            )
            sale_id = cursor.lastrowid

            for product_id, qty, price in selected_items:
                cursor.execute(
                    "INSERT INTO sale_items (sale_id, product_id, quantity, price, tenant) VALUES (?, ?, ?, ?, ?)",
                    (sale_id, product_id, qty, price, TENANT),
                )
                cursor.execute(
                    "UPDATE products SET quantity = quantity - ? WHERE id = ? AND tenant = ?",
                    (qty, product_id, TENANT),
                )

            conn.commit()

            log_action(
                session["user_id"],
                "SALE",
                f"Sale ID {sale_id} recorded, Total: {grand_total}",
            )
            conn.close()

            flash("Sale recorded successfully!", "info")
            return redirect_to_tenant_path("sales-history")
        except sqlite3.Error:
            conn.rollback()
            conn.close()
            flash("Unable to record sale right now.", "error")
            return redirect_to_tenant_path("record-sale")

    conn.close()
    return render_template("record-sale.html", products=products, tenant=TENANT)


# ---------- SALES HISTORY ----------
@app.route(f"/{TENANT}/sales-history")
@login_required
def sales_history():

    try:
        conn = get_db_connection()

        sales = conn.execute(
            "SELECT * FROM sales WHERE tenant = ? ORDER BY date DESC",
            (TENANT,),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to load sales history right now.", "error")
        sales = []

    return render_template(
        "sales-history.html",
        sales=sales,
        tenant=TENANT,
    )


# ---------- INVOICE ----------
@app.route(f"/{TENANT}/invoice/<int:sale_id>")
@login_required
def download_invoice(sale_id):

    try:
        conn = get_db_connection()

        sale = conn.execute(
            "SELECT * FROM sales WHERE id = ? AND tenant = ?",
            (sale_id, TENANT),
        ).fetchone()

        items = conn.execute(
            """
            SELECT p.name, si.quantity, si.price
            FROM sale_items si
            JOIN products p ON si.product_id = p.id AND p.tenant = ?
            WHERE si.sale_id = ? AND si.tenant = ?
            """,
            (TENANT, sale_id, TENANT),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to load that invoice right now.", "error")
        return redirect_to_tenant_path("sales-history")

    if not sale:
        flash("Invoice not found.", "error")
        return redirect_to_tenant_path("sales-history")

    return render_template(
        "invoice.html",
        sale=sale,
        items=items,
        tenant=TENANT,
    )


# ---------- QR ----------
@app.route(f"/{TENANT}/qr/<int:sale_id>")
@login_required
def generate_qr(sale_id):
    try:
        conn = get_db_connection()
        sale = conn.execute(
            "SELECT id FROM sales WHERE id = ? AND tenant = ?",
            (sale_id, TENANT),
        ).fetchone()
        conn.close()
    except sqlite3.Error:
        flash("Unable to generate QR right now.", "error")
        return redirect_to_tenant_path("sales-history")

    # Block invalid or cross-tenant access
    if not sale:
        flash("Invalid or unauthorized invoice.", "error")
        return redirect_to_tenant_path("sales-history")

    # Safe to generate QR
    qr = qrcode.make(f"Invoice ID: {sale_id} | Tenant: {TENANT}")
    img_io = io.BytesIO()
    qr.save(img_io, "PNG")
    img_io.seek(0)

    return send_file(
        img_io,
        mimetype="image/png",
        as_attachment=False,
        download_name=f"qr_{sale_id}.png",
        max_age=0,
    )


@app.route(f"/{TENANT}/inventory-tools")
@login_required
@role_required("admin")
def inventory_tools():
    return render_template("inventory-tools.html", tenant=TENANT)


# ---------- DOWNLOAD INVENTORY ----------
@app.route(f"/{TENANT}/download-inventory")
@login_required
@role_required("admin")
def download_inventory():

    try:
        conn = get_db_connection()

        products = conn.execute(
            "SELECT id, name, quantity, buying_price, price FROM products WHERE tenant = ?",
            (TENANT,),
        ).fetchall()

        conn.close()

    except sqlite3.Error:
        flash("Unable to export inventory right now.", "error")
        return redirect_to_tenant_path("inventory-tools")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["ID", "Item", "Stock", "Buying Price (KES)", "Selling Price (KES)"])

    for product in products:
        ws.append(
            [
                product["id"],
                product["name"],
                product["quantity"],
                product["buying_price"],
                product["price"],
            ]
        )

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f"{TENANT}_inventory.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# ---------- UPLOAD INVENTORY ----------
@app.route(f"/{TENANT}/upload-inventory", methods=["POST"])
@limiter.limit("5 per minute")
@login_required
@role_required("admin")
def upload_inventory():

    # CSRF check
    if not validate_csrf():
        flash("Invalid request.", "error")
        return redirect_to_tenant_path("inventory-tools")

    # Check file exists
    if "file" not in request.files:
        flash("No file uploaded.", "error")
        return redirect_to_tenant_path("inventory-tools")

    file = request.files["file"]

    # Check filename
    if file.filename == "":
        flash("No file selected.", "error")
        return redirect_to_tenant_path("inventory-tools")

    filename = file.filename.lower()

    # Extension check
    if not filename.endswith(".xlsx"):
        flash("Only Excel (.xlsx) files are allowed.", "error")
        return redirect_to_tenant_path("inventory-tools")

    # File size limit (5MB)
    try:
        file.seek(0, 2)
        size = file.tell()
        file.seek(0)
    except Exception:
        flash("File processing error.", "error")
        return redirect_to_tenant_path("inventory-tools")

    if size > 5 * 1024 * 1024:
        flash("File too large (max 5MB).", "error")
        return redirect_to_tenant_path("inventory-tools")

    # Validate Excel content
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
    except Exception:
        flash("Invalid Excel file.", "error")
        return redirect_to_tenant_path("inventory-tools")

    rows_updated = 0

    try:
        conn = get_db_connection()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue

            product_id, name, stock, buying_price, selling_price = row[:5]

            # Validate product_id
            if not isinstance(product_id, int):
                continue

            # Validate name
            if not isinstance(name, str) or not name.strip():
                continue

            # Validate numeric values
            try:
                stock = int(stock)
                buying_price = float(buying_price)
                selling_price = float(selling_price)
            except (TypeError, ValueError):
                continue

            # Business rules
            if stock < 0 or buying_price < 0 or selling_price < 0:
                continue

            # Update DB (tenant-safe)
            conn.execute(
                """
                UPDATE products
                SET name = ?, quantity = ?, buying_price = ?, price = ?
                WHERE id = ? AND tenant = ?
                """,
                (name.strip(), stock, buying_price, selling_price, product_id, TENANT),
            )

            rows_updated += 1

        # Log action
        conn.execute(
            "INSERT INTO logs (user_id, action, details) VALUES (?, ?, ?)",
            (
                session["user_id"],
                "INVENTORY_UPLOAD",
                f"Inventory upload processed: {rows_updated} rows updated",
            ),
        )

        conn.commit()
        conn.close()

    except Exception:
        flash("Unable to update inventory.", "error")
        return redirect_to_tenant_path("inventory-tools")

    # Accurate feedback
    if rows_updated == 0:
        flash("Upload failed. No valid rows found.", "error")
    else:
        flash(f"Inventory updated successfully! ({rows_updated} items)", "success")

    return redirect_to_tenant_path("inventory-tools")


# ---------- ERROR HANDLERS ----------
@app.errorhandler(403)
def forbidden(error):
    flash("You do not have permission to access this page.", "error")
    return render_template("403.html", tenant=TENANT), 403


@app.errorhandler(404)
def not_found(error):
    flash("Page not found.", "error")
    return render_template("404.html", tenant=TENANT), 404


if __name__ == "__main__":
    app.run(debug=False)

